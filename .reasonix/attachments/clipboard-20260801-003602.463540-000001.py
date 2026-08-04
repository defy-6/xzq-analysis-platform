#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
批量调用高德 Web 服务驾车路径规划 API，计算 OD 的时间、距离和过路费。

输入：.xlsx 文件，每行一组 OD。
坐标：支持 WGS84 或高德 GCJ-02；WGS84 会在请求前转换为 GCJ-02。
输出：在原表末尾追加路径规划结果，支持断点续跑。
"""

from __future__ import annotations

import argparse
import json
import math
import os
import random
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


API_URL = "https://restapi.amap.com/v3/direction/driving"
PI = math.pi
A = 6378245.0
EE = 0.00669342162296594323

# 自动识别常见字段。若未识别，可用命令行参数显式指定。
COORDINATE_SETS = [
    ("起点_wgs84_经", "起点_wgs84_纬", "终点_wgs84_经", "终点_wgs84_纬", "wgs84"),
    ("起点wgs84_经", "起点wgs84_纬", "终点wgs84_经", "终点wgs84_纬", "wgs84"),
    ("A_wgs84_经", "A_wgs84_纬", "B_wgs84_经", "B_wgs84_纬", "wgs84"),
    ("O_wgs84_经", "O_wgs84_纬", "D_wgs84_经", "D_wgs84_纬", "wgs84"),
    ("origin_wgs84_lon", "origin_wgs84_lat", "destination_wgs84_lon", "destination_wgs84_lat", "wgs84"),
    ("起点_GCJ02_经", "起点_GCJ02_纬", "终点_GCJ02_经", "终点_GCJ02_纬", "gcj02"),
    ("起点_gcj02_经", "起点_gcj02_纬", "终点_gcj02_经", "终点_gcj02_纬", "gcj02"),
    ("origin_lon", "origin_lat", "destination_lon", "destination_lat", "gcj02"),
]

RESULT_COLUMNS = [
    "驾车距离_米",
    "驾车距离_公里",
    "驾车时间_秒",
    "驾车时间_分钟",
    "过路费_元",
    "收费路段距离_米",
    "红绿灯数",
    "高德驾车策略",
    "实际请求起点_GCJ02",
    "实际请求终点_GCJ02",
    "高德查询状态",
    "高德错误信息",
    "查询时间",
]


def transform_lat(x: float, y: float) -> float:
    ret = -100 + 2 * x + 3 * y + 0.2 * y * y + 0.1 * x * y
    ret += 0.2 * math.sqrt(abs(x))
    ret += (20 * math.sin(6 * x * PI) + 20 * math.sin(2 * x * PI)) * 2 / 3
    ret += (20 * math.sin(y * PI) + 40 * math.sin(y / 3 * PI)) * 2 / 3
    ret += (160 * math.sin(y / 12 * PI) + 320 * math.sin(y * PI / 30)) * 2 / 3
    return ret


def transform_lon(x: float, y: float) -> float:
    ret = 300 + x + 2 * y + 0.1 * x * x + 0.1 * x * y
    ret += 0.1 * math.sqrt(abs(x))
    ret += (20 * math.sin(6 * x * PI) + 20 * math.sin(2 * x * PI)) * 2 / 3
    ret += (20 * math.sin(x * PI) + 40 * math.sin(x / 3 * PI)) * 2 / 3
    ret += (150 * math.sin(x / 12 * PI) + 300 * math.sin(x / 30 * PI)) * 2 / 3
    return ret


def outside_china(lon: float, lat: float) -> bool:
    return not (72.004 <= lon <= 137.8347 and 0.8293 <= lat <= 55.8271)


def wgs84_to_gcj02(lon: float, lat: float) -> tuple[float, float]:
    if outside_china(lon, lat):
        return lon, lat
    dlat = transform_lat(lon - 105, lat - 35)
    dlon = transform_lon(lon - 105, lat - 35)
    radlat = lat / 180 * PI
    magic = math.sin(radlat)
    magic = 1 - EE * magic * magic
    sqrtmagic = math.sqrt(magic)
    dlat = dlat * 180 / ((A * (1 - EE)) / (magic * sqrtmagic) * PI)
    dlon = dlon * 180 / (A / sqrtmagic * math.cos(radlat) * PI)
    return lon + dlon, lat + dlat


def clean_header(value: Any) -> str:
    return str(value).strip().replace("\n", "").replace("\r", "") if value is not None else ""


def find_header_row(ws, specified: int | None) -> int:
    if specified:
        return specified
    expected = {name for item in COORDINATE_SETS for name in item[:4]}
    best_row, best_score = 1, -1
    for row in range(1, min(ws.max_row, 20) + 1):
        values = {clean_header(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        score = len(values & expected)
        if score > best_score:
            best_row, best_score = row, score
    return best_row


def header_map(ws, header_row: int) -> dict[str, int]:
    result = {}
    for col in range(1, ws.max_column + 1):
        name = clean_header(ws.cell(header_row, col).value)
        if name:
            result[name] = col
    return result


def resolve_coordinate_columns(headers: dict[str, int], args):
    explicit = [args.origin_lon, args.origin_lat, args.destination_lon, args.destination_lat]
    if any(explicit):
        if not all(explicit):
            raise ValueError("显式指定坐标字段时，四个字段参数必须全部填写。")
        missing = [name for name in explicit if name not in headers]
        if missing:
            raise ValueError(f"找不到指定字段：{missing}")
        return tuple(headers[name] for name in explicit), args.coord_system

    for o_lon, o_lat, d_lon, d_lat, system in COORDINATE_SETS:
        if all(name in headers for name in (o_lon, o_lat, d_lon, d_lat)):
            return (headers[o_lon], headers[o_lat], headers[d_lon], headers[d_lat]), system
    raise ValueError(
        "无法自动识别四个坐标字段。当前字段为：\n"
        + repr(list(headers))
        + "\n请使用 --origin-lon、--origin-lat、--destination-lon、"
        "--destination-lat 和 --coord-system 显式指定。"
    )


def as_float(value: Any, field: str) -> float:
    if value is None or str(value).strip() == "":
        raise ValueError(f"{field}为空")
    number = float(value)
    if not math.isfinite(number):
        raise ValueError(f"{field}不是有效数字")
    return number


def api_request(
    key: str,
    origin: tuple[float, float],
    destination: tuple[float, float],
    strategy: int,
    timeout: float,
    retries: int,
) -> dict:
    params = {
        "key": key,
        "origin": f"{origin[0]:.6f},{origin[1]:.6f}",
        "destination": f"{destination[0]:.6f},{destination[1]:.6f}",
        "strategy": str(strategy),
        "extensions": "all",
        "output": "json",
    }
    url = API_URL + "?" + urllib.parse.urlencode(params)
    last_error = ""
    for attempt in range(retries + 1):
        try:
            request = urllib.request.Request(url, headers={"User-Agent": "OD-Amap-Batch/1.0"})
            with urllib.request.urlopen(request, timeout=timeout) as response:
                data = json.loads(response.read().decode("utf-8"))
            if str(data.get("status")) != "1":
                raise RuntimeError(
                    f"{data.get('info', 'API失败')} (infocode={data.get('infocode', '')})"
                )
            paths = data.get("route", {}).get("paths") or []
            if not paths:
                raise RuntimeError("API成功返回，但没有可用驾车方案")
            return paths[0]
        except Exception as exc:
            last_error = str(exc)
            if attempt < retries:
                time.sleep(min(2 ** attempt + random.random(), 8))
    raise RuntimeError(last_error)


def first_number(value: Any, default=None):
    if isinstance(value, list):
        value = value[0] if value else default
    if value in (None, "", []):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def parse_args():
    parser = argparse.ArgumentParser(description="批量计算高德驾车时间、距离和过路费")
    parser.add_argument("input", help="输入 .xlsx 文件")
    parser.add_argument("-o", "--output", help="输出文件；默认在输入文件名后加“_驾车结果”")
    parser.add_argument("--key", help="高德 Web 服务 Key；默认读取 AMAP_API_KEY")
    parser.add_argument("--sheet", help="工作表名称；默认使用活动工作表")
    parser.add_argument("--header-row", type=int, help="表头所在行号；默认自动识别")
    parser.add_argument("--origin-lon", help="起点经度字段名")
    parser.add_argument("--origin-lat", help="起点纬度字段名")
    parser.add_argument("--destination-lon", help="终点经度字段名")
    parser.add_argument("--destination-lat", help="终点纬度字段名")
    parser.add_argument("--coord-system", choices=["wgs84", "gcj02"], default="wgs84")
    parser.add_argument("--strategy", type=int, default=0, help="高德驾车策略，默认0")
    parser.add_argument("--delay", type=float, default=0.25, help="请求间隔秒数，默认0.25")
    parser.add_argument("--timeout", type=float, default=20, help="单次请求超时秒数")
    parser.add_argument("--retries", type=int, default=3, help="失败重试次数")
    parser.add_argument("--save-every", type=int, default=20, help="每处理多少行保存一次")
    parser.add_argument("--overwrite", action="store_true", help="重新查询已有成功结果")
    return parser.parse_args()


def main():
    args = parse_args()
    api_key = args.key or os.getenv("AMAP_API_KEY")
    if not api_key:
        sys.exit("未找到 API Key。请设置环境变量 AMAP_API_KEY，或使用 --key。")

    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists():
        sys.exit(f"输入文件不存在：{input_path}")
    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else input_path.with_name(input_path.stem + "_驾车结果.xlsx")
    )

    # 输出文件存在时从输出文件续跑，否则从输入文件开始。
    working_path = output_path if output_path.exists() and not args.overwrite else input_path
    wb = load_workbook(working_path)
    ws = wb[args.sheet] if args.sheet else wb.active
    header_row = find_header_row(ws, args.header_row)
    headers = header_map(ws, header_row)
    coord_cols, detected_system = resolve_coordinate_columns(headers, args)
    coord_system = args.coord_system if any(
        [args.origin_lon, args.origin_lat, args.destination_lon, args.destination_lat]
    ) else detected_system

    result_cols = {}
    for name in RESULT_COLUMNS:
        if name in headers:
            result_cols[name] = headers[name]
        else:
            col = ws.max_column + 1
            ws.cell(header_row, col, name)
            headers[name] = col
            result_cols[name] = col

    print(f"工作表：{ws.title}；表头行：{header_row}；坐标系：{coord_system}")
    print(f"坐标字段列：{coord_cols}；驾车策略：{args.strategy}")

    success = failed = skipped = 0
    for row in range(header_row + 1, ws.max_row + 1):
        status_cell = ws.cell(row, result_cols["高德查询状态"])
        if not args.overwrite and status_cell.value == "成功":
            skipped += 1
            continue
        try:
            o_lon = as_float(ws.cell(row, coord_cols[0]).value, "起点经度")
            o_lat = as_float(ws.cell(row, coord_cols[1]).value, "起点纬度")
            d_lon = as_float(ws.cell(row, coord_cols[2]).value, "终点经度")
            d_lat = as_float(ws.cell(row, coord_cols[3]).value, "终点纬度")
            if not (-180 <= o_lon <= 180 and -90 <= o_lat <= 90):
                raise ValueError("起点经纬度超出有效范围")
            if not (-180 <= d_lon <= 180 and -90 <= d_lat <= 90):
                raise ValueError("终点经纬度超出有效范围")

            if coord_system == "wgs84":
                origin = wgs84_to_gcj02(o_lon, o_lat)
                destination = wgs84_to_gcj02(d_lon, d_lat)
            else:
                origin, destination = (o_lon, o_lat), (d_lon, d_lat)

            path = api_request(
                api_key, origin, destination, args.strategy, args.timeout, args.retries
            )
            distance = first_number(path.get("distance"))
            duration = first_number(path.get("duration"))
            tolls = first_number(path.get("tolls"), 0)
            toll_distance = first_number(path.get("toll_distance"), 0)
            traffic_lights = first_number(path.get("traffic_lights"))

            values = {
                "驾车距离_米": distance,
                "驾车距离_公里": distance / 1000 if distance is not None else None,
                "驾车时间_秒": duration,
                "驾车时间_分钟": duration / 60 if duration is not None else None,
                "过路费_元": tolls,
                "收费路段距离_米": toll_distance,
                "红绿灯数": traffic_lights,
                "高德驾车策略": args.strategy,
                "实际请求起点_GCJ02": f"{origin[0]:.6f},{origin[1]:.6f}",
                "实际请求终点_GCJ02": f"{destination[0]:.6f},{destination[1]:.6f}",
                "高德查询状态": "成功",
                "高德错误信息": None,
                "查询时间": time.strftime("%Y-%m-%d %H:%M:%S"),
            }
            success += 1
        except Exception as exc:
            values = {
                "高德查询状态": "失败",
                "高德错误信息": str(exc),
                "高德驾车策略": args.strategy,
                "查询时间": time.strftime("%Y-%m-%d %H:%M:%S"),
            }
            failed += 1

        for name, value in values.items():
            ws.cell(row, result_cols[name], value)

        processed = success + failed
        if processed % args.save_every == 0:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            print(f"已处理 {processed} 条：成功 {success}，失败 {failed}，跳过 {skipped}")
        time.sleep(max(args.delay, 0))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"完成：成功 {success}，失败 {failed}，跳过 {skipped}")
    print(f"输出：{output_path}")


if __name__ == "__main__":
    main()
