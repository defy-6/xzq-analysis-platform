from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PLATFORM_ROOT = Path(__file__).resolve().parents[3]
SOURCE = PLATFORM_ROOT / "data" / "raw" / "transport" / "driving" / "county-government" / "厦漳泉区县政府所在地_OD.xlsx"
POPULATION_INTERFACE = Path(__file__).resolve().parents[1] / "public" / "data" / "population-flow.json"
OUT = Path(__file__).resolve().parents[1] / "public" / "data" / "transport-accessibility.json"


def avg(values):
    values = [value for value in values if value is not None]
    return sum(values) / len(values) if values else 0.0


def main():
    population = json.loads(POPULATION_INTERFACE.read_text(encoding="utf-8"))
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = {name: position for position, name in enumerate(headers)}
    required = [
        "O_cityname", "O_adname", "O_wgs84_经", "O_wgs84_纬",
        "D_cityname", "D_adname", "D_wgs84_经", "D_wgs84_纬",
        "耗时", "里程（米）", "过路费（元）",
    ]
    missing = [name for name in required if name not in index]
    if missing:
        raise ValueError(f"缺少字段: {missing}")

    records = []
    centers = {}
    by_direction = {}
    node_outbound = defaultdict(list)
    for values in sheet.iter_rows(min_row=2, values_only=True):
        oc, county_o = str(values[index["O_cityname"]] or "").strip(), str(values[index["O_adname"]] or "").strip()
        dc, county_d = str(values[index["D_cityname"]] or "").strip(), str(values[index["D_adname"]] or "").strip()
        if not oc or not county_o or not dc or not county_d or (oc == dc and county_o == county_d):
            continue
        duration_seconds = float(values[index["耗时"]] or 0)
        distance_meters = float(values[index["里程（米）"]] or 0)
        toll = float(values[index["过路费（元）"]] or 0)
        ox, oy = float(values[index["O_wgs84_经"]]), float(values[index["O_wgs84_纬"]])
        dx, dy = float(values[index["D_wgs84_经"]]), float(values[index["D_wgs84_纬"]])
        item = [oc, county_o, dc, county_d, duration_seconds, duration_seconds / 60, distance_meters, distance_meters / 1000, toll]
        records.append(item)
        by_direction[(oc, county_o, dc, county_d)] = item
        node_outbound[(oc, county_o)].append(item)
        centers[f"{oc}|{county_o}"] = [ox, oy]
        centers[f"{dc}|{county_d}"] = [dx, dy]
    workbook.close()

    nodes = sorted(node_outbound)
    pair_records = []
    incomplete_pairs = []
    for left_pos, left in enumerate(nodes):
        for right in nodes[left_pos + 1:]:
            forward = by_direction.get((*left, *right))
            backward = by_direction.get((*right, *left))
            if not forward or not backward:
                incomplete_pairs.append([*left, *right])
                continue
            pair_records.append([
                *left, *right,
                avg([forward[5], backward[5]]),
                avg([forward[7], backward[7]]),
                avg([forward[8], backward[8]]),
                forward[5], backward[5], forward[7], backward[7], forward[8], backward[8],
            ])

    node_stats = []
    for city, county in nodes:
        items = node_outbound[(city, county)]
        node_stats.append({
            "city": city,
            "county": county,
            "avgTime": avg([row[5] for row in items]),
            "avgDistance": avg([row[7] for row in items]),
            "avgToll": avg([row[8] for row in items]),
            "coverage90": sum(row[5] <= 90 for row in items),
            "coverage120": sum(row[5] <= 120 for row in items),
        })
    node_stats.sort(key=lambda item: item["avgTime"])
    for rank, item in enumerate(node_stats, start=1):
        item["rank"] = rank

    city_groups = defaultdict(list)
    for row in records:
        city_groups[(row[0], row[2])].append(row)
    city_pair_stats = [{
        "originCity": key[0],
        "destinationCity": key[1],
        "avgTime": avg([row[5] for row in items]),
        "avgDistance": avg([row[7] for row in items]),
        "avgToll": avg([row[8] for row in items]),
        "odCount": len(items),
    } for key, items in sorted(city_groups.items())]

    payload = {
        "meta": {
            "source": SOURCE.name,
            "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
            "nodeCount": len(nodes),
            "directedOdCount": len(records),
            "pairCount": len(pair_records),
            "incompletePairCount": len(incomplete_pairs),
            "durationUnit": "分钟",
            "distanceUnit": "公里",
            "tollUnit": "元",
            "distanceDefinition": "无向区县对距离取A→B与B→A区县政府驻地驾车里程的算术平均值",
        },
        "governmentCenters": centers,
        "countyBoundaries": population["countyBoundaries"],
        "records": records,
        "pairRecords": pair_records,
        "nodeStats": node_stats,
        "cityPairStats": city_pair_stats,
        "incompletePairs": incomplete_pairs,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(json.dumps({"output": str(OUT), **payload["meta"]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
