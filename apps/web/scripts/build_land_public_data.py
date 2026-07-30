from __future__ import annotations

import json
import csv
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PLATFORM_ROOT = Path(__file__).resolve().parents[3]
WEB_ROOT = Path(__file__).resolve().parents[1]
LAND_SOURCE = PLATFORM_ROOT / "data" / "raw" / "land-use" / "厦漳泉用地数据_整理成果.xlsx"
POI_DETAIL_SOURCE = PLATFORM_ROOT / "data" / "raw" / "public-services" / "poi" / "厦漳泉POI.csv"
DEMOGRAPHIC_SOURCE = PLATFORM_ROOT / "data" / "raw" / "demographics" / "厦漳泉都市圈区县常住人口与GDP.xlsx"
BOUNDARY_SOURCE = WEB_ROOT / "public" / "data" / "population-flow.json"
LAND_OUT = WEB_ROOT / "public" / "data" / "land-use.json"
POI_OUT = WEB_ROOT / "public" / "data" / "public-services.json"


def rows_as_dicts(sheet):
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(value or "") for value in next(iterator)]
    for values in iterator:
        if not any(value is not None for value in values):
            continue
        yield {headers[index]: value for index, value in enumerate(values)}


def number(value):
    if value is None:
        return 0.0
    return float(str(value).replace("\xa0", "").replace(",", "").strip())


EXCLUDED_POI_MAJORS = {"地名地址信息", "通行设施", "虚拟数据", "道路附属设施"}
POI_FUNCTIONS = ["商业消费", "餐饮休闲", "产业就业", "物流服务", "公共管理", "教育文化", "医疗健康", "交通门户", "文旅生活"]


def poi_function(major, middle, minor, name):
    if major in EXCLUDED_POI_MAJORS:
        return ""
    combined = f"{major}|{middle}|{minor}|{name}"
    if any(word in combined for word in ("物流", "仓储", "快递", "货运", "配送", "货物集散")):
        return "物流服务"
    if major in {"购物服务", "金融保险服务", "汽车销售"}:
        return "商业消费"
    if major in {"餐饮服务", "体育休闲服务"}:
        return "餐饮休闲"
    if major in {"公司企业", "商务住宅"}:
        return "产业就业"
    if major in {"政府机构及社会团体", "公共设施"}:
        return "公共管理"
    if major == "科教文化服务":
        return "教育文化"
    if major == "医疗保健服务":
        return "医疗健康"
    if major == "交通设施服务":
        return "交通门户"
    return "文旅生活"


def main():
    spatial = json.loads(BOUNDARY_SOURCE.read_text(encoding="utf-8"))
    boundary_features = spatial["countyBoundaries"]
    valid_counties = {feature["properties"]["name"] for feature in boundary_features["features"]}
    city_by_county = {feature["properties"]["name"]: feature["properties"]["city"] for feature in boundary_features["features"]}

    land_book = load_workbook(LAND_SOURCE, read_only=True, data_only=True)
    metric_rows = list(rows_as_dicts(land_book["区县指标"]))
    lq_rows = list(rows_as_dicts(land_book["区县LQ"]))
    three_rows = list(rows_as_dicts(land_book["区县三类汇总"]))
    ownership_rows = list(rows_as_dicts(land_book["权属交叉"]))
    lq_by_county = defaultdict(dict)
    for row in lq_rows:
        lq_by_county[str(row["区县"])][str(row["功能类别"])] = float(row["区位熵LQ"] or 0)
    three_by_county = {str(row["区县"]): row for row in three_rows}
    ownership_by_county = {str(row["区县"]): row for row in ownership_rows}
    land_records = []
    for row in metric_rows:
        county = str(row["区县"])
        three = three_by_county[county]
        ownership = ownership_by_county[county]
        land_records.append({
            "city": str(row["城市"]),
            "county": county,
            "totalArea": float(row["行政区总面积_公顷"] or 0),
            "agriculturalArea": float(three["农用地_公顷"] or 0),
            "constructionArea": float(row["建设用地_公顷"] or 0),
            "unusedArea": float(three["未利用地_公顷"] or 0),
            "developmentIntensity": float(row["开发强度"] or 0),
            "agriculturalShare": float(row["农用地占比"] or 0),
            "unusedShare": float(row["未利用地占比"] or 0),
            "urbanHousingShare": float(row["城镇住宅占建设用地"] or 0),
            "ruralHousingShare": float(row["农村宅基地占建设用地"] or 0),
            "urbanRuralHousingRatio": float(row["城乡居住结构比"] or 0),
            "urbanHousingArea": float(row["城镇住宅_公顷"] or 0),
            "ruralHousingArea": float(row["农村宅基地_公顷"] or 0),
            "industrialArea": float(row["工业_公顷"] or 0),
            "logisticsArea": float(row["物流仓储_公顷"] or 0),
            "industryWarehouseShare": float(row["工业及仓储占建设用地"] or 0),
            "industrialDevelopmentDensity": float(row["产业开发密度"] or 0),
            "commercialShare": float(row["商业服务占建设用地"] or 0),
            "commercialArea": float(row["商业服务_公顷"] or 0),
            "publicServiceShare": float(row["公共服务占建设用地"] or 0),
            "publicServiceArea": float(row["公共服务_公顷"] or 0),
            "centralFunctionIndex": float(row["中心功能指数"] or 0),
            "transportGatewayIndex": float(row["交通门户指数"] or 0),
            "gatewayArea": float(row["交通门户_公顷"] or 0),
            "roadArea": float(row["公路_公顷"] or 0),
            "roadShare": float(row["公路占建设用地"] or 0),
            "portShare": float(row["港口码头占建设用地"] or 0),
            "portArea": float(row["港口码头_公顷"] or 0),
            "miningArea": float(row["采矿_公顷"] or 0),
            "publicSpaceShare": float(row["公共空间占建设用地"] or 0),
            "publicSpaceArea": float(row["公共空间_公顷"] or 0),
            "landUseMix": float(row["建设用地混合度"] or 0),
            "scope201Construction": float(ownership["201_城市范围_建设用地"] or 0),
            "scope202Construction": float(ownership["202_建制镇范围_建设用地"] or 0),
            "scope203Construction": float(ownership["203_村庄范围_建设用地"] or 0),
            "scope201AIndustrial": float(ownership["201A_城市独立工业用地_建设用地"] or 0),
            "scope202AIndustrial": float(ownership["202A_建制镇独立工业用地_建设用地"] or 0),
            "scope203AIndustrial": float(ownership["203A_村庄独立工业用地_建设用地"] or 0),
            "lq": lq_by_county[county],
        })
    land_book.close()
    if len(land_records) != 28 or {row["county"] for row in land_records} != valid_counties:
        raise ValueError("用地区县与网站28区县边界不一致")
    land_payload = {
        "meta": {
            "source": LAND_SOURCE.name,
            "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
            "countyCount": len(land_records),
            "year": "未提供（单期横截面）",
            "areaUnit": "公顷",
            "note": "行政区总面积按农用地、建设用地、未利用地面积合计；未利用地不能直接解释为可开发后备土地。",
        },
        "countyBoundaries": boundary_features,
        "records": land_records,
    }

    demographic_book = load_workbook(DEMOGRAPHIC_SOURCE, read_only=True, data_only=True)
    demographic_rows = list(rows_as_dicts(demographic_book.active))
    demographic_book.close()
    population_by_county = {
        str(row["区县"]).strip(): number(row["2024年常住人口（万人）"])
        for row in demographic_rows
    }
    if set(population_by_county) != valid_counties:
        missing = sorted(valid_counties - set(population_by_county))
        extra = sorted(set(population_by_county) - valid_counties)
        raise ValueError(f"常住人口区县口径不一致：missing={missing}, extra={extra}")
    construction_by_county = {row["county"]: row["constructionArea"] for row in land_records}

    rename = {"龙海市": "龙海区", "长泰县": "长泰区"}
    counts = defaultdict(int)
    functional_counts = defaultdict(int)
    seen_functional = set()
    excluded_functional_rows = 0
    duplicate_functional_rows = 0
    raw_names = set()
    excluded_names = set()
    with POI_DETAIL_SOURCE.open(encoding="utf-8-sig", newline="") as source:
        for row in csv.DictReader(source):
            raw_county = str(row.get("adname") or "").strip()
            raw_names.add(raw_county)
            county = rename.get(raw_county, raw_county)
            if county not in valid_counties:
                excluded_names.add(raw_county)
                continue
            major = str(row.get("大类") or "未分类").strip() or "未分类"
            middle = str(row.get("中类") or "未分类").strip() or "未分类"
            minor = str(row.get("小类") or "未分类").strip() or "未分类"
            counts[(county, major, middle, minor)] += 1
            function = poi_function(major, middle, minor, str(row.get("name") or "").strip())
            if not function:
                excluded_functional_rows += 1
                continue
            dedup_key = (
                county,
                str(row.get("name") or "").strip().lower(),
                str(row.get("typecode") or "").strip(),
                str(row.get("lon_wgs84") or "").strip(),
                str(row.get("lat_wgs84") or "").strip(),
            )
            if dedup_key in seen_functional:
                duplicate_functional_rows += 1
                continue
            seen_functional.add(dedup_key)
            functional_counts[(county, function)] += 1
    major_totals = defaultdict(int)
    middle_totals = defaultdict(int)
    minor_totals = defaultdict(int)
    poi_records = []
    county_totals = defaultdict(int)
    for (county, major, middle, minor), value in counts.items():
        poi_records.append([city_by_county[county], county, major, middle, minor, value])
        county_totals[county] += value
        major_totals[major] += value
        middle_totals[(major, middle)] += value
        minor_totals[(major, middle, minor)] += value
    category_tree = []
    for major, major_total in sorted(major_totals.items(), key=lambda item: (-item[1], item[0])):
        middle_children = []
        for (parent_major, middle), middle_total in sorted(middle_totals.items(), key=lambda item: (-item[1], item[0][1])):
            if parent_major != major:
                continue
            minor_children = [
                {"name": minor, "count": total}
                for (minor_major, minor_middle, minor), total in sorted(minor_totals.items(), key=lambda item: (-item[1], item[0][2]))
                if minor_major == major and minor_middle == middle
            ]
            middle_children.append({"name": middle, "count": middle_total, "children": minor_children})
        category_tree.append({"name": major, "count": major_total, "children": middle_children})
    poi_records.sort(key=lambda row: (row[1], row[2], row[3], row[4]))
    poi_payload = {
        "meta": {
            "source": POI_DETAIL_SOURCE.name,
            "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
            "countyCount": len(valid_counties),
            "categoryCount": len(minor_totals),
            "majorCategoryCount": len(major_totals),
            "middleCategoryCount": len(middle_totals),
            "minorCategoryCount": len(minor_totals),
            "poiTotal": sum(major_totals.values()),
            "rawCountyNameCount": len(raw_names),
            "renamedCounties": rename,
            "excludedCountyNames": sorted(excluded_names),
            "rule": "按现行区县名称合并龙海市/龙海区、长泰县/长泰区，并仅保留厦门、漳州、泉州三市28个区县。",
            "populationSource": DEMOGRAPHIC_SOURCE.name,
            "populationYear": "2024",
            "landSource": LAND_SOURCE.name,
            "normalizationRule": "人均供给按每万常住人POI数计算；用地密度按每平方公里建设用地POI数计算（1平方公里=100公顷）。",
            "functionalCategories": POI_FUNCTIONS,
            "functionalPoiTotal": sum(functional_counts.values()),
            "excludedFunctionalRows": excluded_functional_rows,
            "duplicateFunctionalRows": duplicate_functional_rows,
            "functionalRule": "排除地名地址信息、通行设施、虚拟数据和道路附属设施；有效记录按区县＋名称＋类型代码＋WGS84坐标去重。",
        },
        "countyBoundaries": boundary_features,
        "categoryTree": category_tree,
        "countyTotals": dict(county_totals),
        "countyContext": {
            county: {
                "residentPopulationWan": population_by_county[county],
                "constructionAreaHa": construction_by_county[county],
                "commercialAreaHa": next(row["commercialArea"] for row in land_records if row["county"] == county),
                "industryLogisticsAreaHa": next((row["industrialArea"] + row["logisticsArea"]) for row in land_records if row["county"] == county),
                "publicServiceAreaHa": next(row["publicServiceArea"] for row in land_records if row["county"] == county),
                "gatewayAreaHa": next(row["gatewayArea"] for row in land_records if row["county"] == county),
            }
            for county in sorted(valid_counties)
        },
        "functionalRecords": [
            [city_by_county[county], county, function, value]
            for (county, function), value in sorted(functional_counts.items())
        ],
        "records": poi_records,
    }
    LAND_OUT.write_text(json.dumps(land_payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    POI_OUT.write_text(json.dumps(poi_payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(json.dumps({
        "land": {"output": str(LAND_OUT), **land_payload["meta"]},
        "publicServices": {"output": str(POI_OUT), **poi_payload["meta"]},
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
