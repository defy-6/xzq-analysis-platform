from __future__ import annotations

import json
import sqlite3
import struct
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PLATFORM_ROOT = Path(__file__).resolve().parents[4]
POPULATION_DIR = PLATFORM_ROOT / "data" / "raw" / "population"
SOURCE_FILE = POPULATION_DIR / "厦漳泉乡镇级人口流动_汇总版.xlsx"
GPKG = PLATFORM_ROOT / "data" / "raw" / "spatial" / "fujian.gpkg"
OUT = Path(__file__).resolve().parents[2] / "public" / "data" / "population-flow.json"
# 高德 API 批量获取的镇街政府驻地经纬度（WGS84）OD 表，作为镇街中心点最优先来源
AMAP_COORDS_FILE = PLATFORM_ROOT / "data" / "raw" / "transport" / "driving" / "township-government" / "厦漳泉乡镇街政府经纬度_WGS84_OD_驾车结果.xlsx"
CITY_CODES = {"3502": "厦门市", "3505": "泉州市", "3506": "漳州市"}
CITIES = set(CITY_CODES.values())


COUNTIES = (
    "思明区", "湖里区", "集美区", "海沧区", "同安区", "翔安区",
    "芗城区", "龙文区", "龙海区", "长泰区", "漳浦县", "云霄县", "东山县", "诏安县", "南靖县", "平和县", "华安县",
    "鲤城区", "丰泽区", "洛江区", "泉港区", "石狮市", "晋江市", "南安市", "惠安县", "安溪县", "永春县", "德化县",
)


def text(value):
    return "" if value is None else str(value).strip()


def normalize_town(value):
    name = text(value)
    # 去掉镇街名开头的区县前缀，如 "芗城区通北街道" → "通北街道"（与前端 mapDisplayName 一致）
    for county in COUNTIES:
        if name.startswith(county):
            name = name[len(county):]
            break
    for suffix in ("街道办事处", "街道", "民族乡", "畲族乡", "镇", "乡"):
        if name.endswith(suffix):
            return name[: -len(suffix)]
    return name


def polygon_centroid(polygons):
    """多个多边形（含内环）的面积加权质心；退化时返回 None。"""
    total_x = total_y = total_area = 0.0
    for polygon in polygons:
        for ring in polygon:
            ring_area = 0.0
            ring_x = ring_y = 0.0
            for i in range(len(ring) - 1):
                x1, y1 = ring[i]
                x2, y2 = ring[i + 1]
                cross = x1 * y2 - x2 * y1
                ring_area += cross
                ring_x += (x1 + x2) * cross
                ring_y += (y1 + y2) * cross
            ring_area /= 2.0
            if abs(ring_area) < 1e-12:
                continue
            abs_area = abs(ring_area)
            total_x += ring_x / (6.0 * ring_area) * abs_area
            total_y += ring_y / (6.0 * ring_area) * abs_area
            total_area += abs_area
    if total_area <= 0:
        return None
    return [total_x / total_area, total_y / total_area]


def load_amap_town_centers():
    """从高德 OD 驾车结果表读取镇街政府驻地坐标（WGS84），键为 区县|~归一化镇街名。"""
    centers = {}
    try:
        workbook = load_workbook(AMAP_COORDS_FILE, read_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        header = [text(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        columns = {name: index for index, name in enumerate(header)}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            for side in ("O", "D"):
                city = row[columns.get(side + "_城市名称", -1)]
                county = row[columns.get(side + "_区县名称", -1)]
                town = row[columns.get(side + "_乡镇街名称", -1)]
                lon = row[columns.get(side + "_wgs84_经", -1)]
                lat = row[columns.get(side + "_wgs84_纬", -1)]
                if city and county and town and lon is not None and lat is not None:
                    key = f"{county}|~{normalize_town(town)}"
                    centers.setdefault(key, [float(lon), float(lat)])
        workbook.close()
    except Exception:
        pass
    return centers


def gpkg_wkb(blob):
    flags = blob[3]
    envelope = (flags >> 1) & 7
    sizes = {0: 0, 1: 32, 2: 48, 3: 48, 4: 64}
    return memoryview(blob)[8 + sizes[envelope] :]


def parse_geometry(data):
    position = 0

    def read():
        nonlocal position
        endian = "<" if data[position] == 1 else ">"
        position += 1
        geometry_type = struct.unpack_from(endian + "I", data, position)[0] % 1000
        position += 4
        if geometry_type == 3:
            ring_count = struct.unpack_from(endian + "I", data, position)[0]
            position += 4
            rings = []
            for _ in range(ring_count):
                point_count = struct.unpack_from(endian + "I", data, position)[0]
                position += 4
                ring = []
                for _ in range(point_count):
                    x, y = struct.unpack_from(endian + "dd", data, position)
                    position += 16
                    ring.append([x, y])
                rings.append(ring)
            return [rings]
        if geometry_type == 6:
            polygon_count = struct.unpack_from(endian + "I", data, position)[0]
            position += 4
            polygons = []
            for _ in range(polygon_count):
                polygons.extend(read())
            return polygons
        raise ValueError(f"Unsupported geometry type: {geometry_type}")

    return read()


def simplify(ring, tolerance=0.00008):
    """Douglas-Peucker simplification in geographic degrees (about 8 m locally).

    The former fixed point sampling kept only every Nth vertex. It distorted
    coastlines and produced visibly angular county boundaries. This method keeps
    vertices where the line actually bends, so the same file size retains much
    more of the source geometry's visual character.
    """
    if len(ring) < 4:
        return ring
    closed = ring[0] == ring[-1]
    points = ring[:-1] if closed else ring
    if len(points) < 3:
        return ring
    keep = {0, len(points) - 1}
    stack = [(0, len(points) - 1)]
    tolerance_sq = tolerance * tolerance
    while stack:
        start, end = stack.pop()
        ax, ay = points[start]
        bx, by = points[end]
        dx, dy = bx - ax, by - ay
        length_sq = dx * dx + dy * dy
        furthest = -1
        max_distance_sq = 0.0
        for index in range(start + 1, end):
            px, py = points[index]
            if length_sq:
                cross = (px - ax) * dy - (py - ay) * dx
                distance_sq = cross * cross / length_sq
            else:
                distance_sq = (px - ax) ** 2 + (py - ay) ** 2
            if distance_sq > max_distance_sq:
                furthest, max_distance_sq = index, distance_sq
        if furthest >= 0 and max_distance_sq > tolerance_sq:
            keep.add(furthest)
            stack.extend(((start, furthest), (furthest, end)))
    result = [points[index] for index in sorted(keep)]
    if closed:
        result.append(result[0])
    return result


def bounds(polygons):
    points = [point for polygon in polygons for ring in polygon for point in ring]
    xs = [point[0] for point in points]
    ys = [point[1] for point in points]
    return min(xs), min(ys), max(xs), max(ys)


def load_boundaries(connection):
    towns = []
    query = """
        select [城市名称], [区县名称], [乡镇街名称], Shape_Area, geom
        from [厦漳泉乡镇街边界]
    """
    for city, county, town, area, blob in connection.execute(query):
        polygons = parse_geometry(gpkg_wkb(blob))
        centroid = polygon_centroid(polygons)
        if centroid is None:
            bbox = bounds(polygons)
            centroid = [(bbox[0] + bbox[2]) / 2, (bbox[1] + bbox[3]) / 2]
        towns.append(
            {
                "city": text(city),
                "county": text(county),
                "town": text(town),
                "center": centroid,
            }
        )

    county_features = []
    county_centers = {}
    county_query = "select XZQDM, XZQMC, geom from [县级调查界限]"
    for code, name, blob in connection.execute(county_query):
        code = text(code)
        if code[:4] not in CITY_CODES:
            continue
        polygons = parse_geometry(gpkg_wkb(blob))
        bbox = bounds(polygons)
        county_name = text(name)
        city = CITY_CODES[code[:4]]
        county_centers[f"{city}|{county_name}"] = [(bbox[0] + bbox[2]) / 2, (bbox[1] + bbox[3]) / 2]
        county_features.append(
            {
                "type": "Feature",
                "properties": {"city": city, "name": county_name, "code": code},
                "geometry": {
                    "type": "MultiPolygon",
                    "coordinates": [[simplify(ring) for ring in polygon] for polygon in polygons],
                },
            }
        )
    return towns, county_features, county_centers


def main():
    with sqlite3.connect(GPKG) as connection:
        towns, county_features, county_centers = load_boundaries(connection)

    workbook = load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    column = {name: index for index, name in enumerate(headers)}

    county_aggregate = defaultdict(lambda: [0, 0])
    town_aggregate = defaultdict(lambda: [0, 0])
    town_centers = {}
    boundary_town_centers = {}
    normalized_boundary_town_centers = {}
    amap_centers = load_amap_town_centers()
    for town in towns:
        exact_key = f'{town["city"]}|{town["county"]}|{town["town"]}'
        normalized_key = f'{town["city"]}|{town["county"]}|~{normalize_town(town["town"])}'
        boundary_town_centers[exact_key] = town["center"]
        normalized_boundary_town_centers.setdefault(normalized_key, town["center"])
    raw_rows = positive_rows = valid_rows = 0
    excluded_nonpositive = excluded_incomplete = 0
    excluded_same_town = excluded_same_town_population = 0
    fallback_town_names = set()
    within_county_population = 0

    endpoints = (
        ("起点地市名称", "起点区县名称", "起点乡镇名称"),
        ("终到地市名称", "终到区县名称", "终到乡镇名称"),
    )

    for row in rows:
        raw_rows += 1
        population = int(row[column["人口数量"]] or 0)
        if population <= 0:
            excluded_nonpositive += 1
            continue
        positive_rows += 1
        source_endpoints = [
            (text(row[column[city_column]]), text(row[column[county_column]]), text(row[column[town_column]]))
            for city_column, county_column, town_column in endpoints
        ]
        if any(city not in CITIES or not county or not town for city, county, town in source_endpoints):
            excluded_incomplete += 1
            continue
        origin_tuple, destination_tuple = source_endpoints
        if origin_tuple == destination_tuple:
            excluded_same_town += 1
            excluded_same_town_population += population
            continue
        valid_rows += 1
        for city, county, town in source_endpoints:
            key = f"{city}|{county}|{town}"
            if key in town_centers:
                continue
            normalized_key = f"{city}|{county}|~{normalize_town(town)}"
            amap_key = f"{county}|~{normalize_town(town)}"
            center = amap_centers.get(amap_key) or boundary_town_centers.get(key) or normalized_boundary_town_centers.get(normalized_key)
            if center is None:
                center = county_centers.get(f"{city}|{county}")
                fallback_town_names.add(key)
            if center is not None:
                town_centers[key] = center

        town_key = (*origin_tuple, *destination_tuple)
        town_aggregate[town_key][0] += population
        town_aggregate[town_key][1] += 1

        if origin_tuple[:2] == destination_tuple[:2]:
            within_county_population += population
            continue
        county_key = (*origin_tuple[:2], *destination_tuple[:2])
        county_aggregate[county_key][0] += population
        county_aggregate[county_key][1] += 1

    workbook.close()

    county_records = [
        [*key, values[0], values[1]] for key, values in county_aggregate.items() if values[0] > 0
    ]
    town_records = [
        [*key, values[0], values[1]] for key, values in town_aggregate.items() if values[0] > 0
    ]
    county_records.sort(key=lambda item: item[4], reverse=True)
    town_records.sort(key=lambda item: item[6], reverse=True)

    payload = {
        "meta": {
            "source": SOURCE_FILE.name,
            "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
            "unit": "人",
            "rawRows": raw_rows,
            "positiveRows": positive_rows,
            "validRows": valid_rows,
            "excludedNonPositive": excluded_nonpositive,
            "excludedIncomplete": excluded_incomplete,
            "excludedSameTown": excluded_same_town,
            "excludedSameTownPopulation": excluded_same_town_population,
            "fallbackTownNames": len(fallback_town_names),
            "withinCountyPopulation": within_county_population,
            "rule": "行政区代码不参与判断，地市、区县和镇街均直接使用源文件名称；排除人口数量不大于0、两端名称不完整或O/D为同一镇街的记录。",
        },
        "countyCenters": county_centers,
        "townCenters": town_centers,
        "countyBoundaries": {"type": "FeatureCollection", "features": county_features},
        "countyRecords": county_records,
        "townRecords": town_records,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(
        json.dumps(
            {
                "output": str(OUT),
                "size": OUT.stat().st_size,
                "countyRecords": len(county_records),
                "townRecords": len(town_records),
                **payload["meta"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
