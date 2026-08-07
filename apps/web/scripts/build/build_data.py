from __future__ import annotations

import json, math, re, sqlite3, struct
from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from openpyxl import load_workbook

PLATFORM_ROOT = Path(__file__).resolve().parents[4]
SOURCE = PLATFORM_ROOT / "data" / "raw" / "enterprise"
GPKG = PLATFORM_ROOT / "data" / "raw" / "spatial" / "fujian.gpkg"
OUT = Path(__file__).resolve().parents[2] / "public" / "data"
CITIES = {"厦门市", "漳州市", "泉州市"}
CITY_CODES = {"3502": "厦门市", "3505": "泉州市", "3506": "漳州市"}
COUNTY_UPDATES={"龙海市":"龙海区","长泰县":"长泰区"}
FX = {"":1,"CNY":1,"USD":6.7684,"HKD":.8632,"EUR":7.7241,"DEM":7.7241/1.95583}
INDUSTRY_CODE_FILE=PLATFORM_ROOT/"data"/"reference"/"industry"/"行业代码表db_code_t_industry_code.xlsx"
INDUSTRY_CATALOG={1:{},2:{},3:{}}

def town_pair_key(record):
    return "↔".join(sorted((f"{record[4]}|{record[5]}",f"{record[7]}|{record[8]}")))

def write_town_endpoints(records):
    endpoint_dir=OUT/"township";endpoint_dir.mkdir(parents=True,exist_ok=True)
    for old in endpoint_dir.glob("pair-*.json"):old.unlink()
    grouped=defaultdict(list)
    for record in records:grouped[town_pair_key(record)].append(record)
    pairs={}
    meta={"amountUnit":"万元人民币","onlyCompleteTownOd":True,"patentUndirected":True,"fields":["relation","industryLevel","industryCode","industryName","oCity","oCounty","oTown","dCity","dCounty","dTown","count","amount","amountCount","uniquePatents","enterprisePairs","maxPairPatents","maxPairShare","unmatched"]}
    for index,key in enumerate(sorted(grouped),1):
        filename=f"pair-{index:03d}.json";pairs[key]=filename
        (endpoint_dir/filename).write_text(json.dumps({"meta":meta,"records":grouped[key]},ensure_ascii=False,separators=(",",":")),encoding="utf-8")
    (OUT/"township-index.json").write_text(json.dumps({"meta":{"pairCount":len(pairs),"recordCount":len(records)},"pairs":pairs},ensure_ascii=False,separators=(",",":")),encoding="utf-8")
    combined=OUT/"township-relations.json"
    combined.write_text(json.dumps({"meta":meta,"records":records},ensure_ascii=False,separators=(",",":")),encoding="utf-8")
    return len(pairs)

def append_dev_log(title, detail, event_type="data"):
    log_file=OUT/"dev-log.json";now=datetime.now().astimezone().isoformat(timespec="seconds")
    try:payload=json.loads(log_file.read_text(encoding="utf-8"))
    except (FileNotFoundError,json.JSONDecodeError):payload={"updatedAt":now,"entries":[]}
    payload["entries"]=(payload.get("entries",[])+[{"time":now,"type":event_type,"title":title,"detail":detail}])[-100:]
    payload["updatedAt"]=now;log_file.write_text(json.dumps(payload,ensure_ascii=False,indent=2),encoding="utf-8")
    markdown=Path(__file__).resolve().parents[2]/"开发日志.md"
    with markdown.open("a",encoding="utf-8") as fh:fh.write(f"\n- {now}｜{title}：{detail}\n")

def txt(v): return "" if v is None else str(v).strip()
def normalized_patent_name(v):
    # 专利名称是去重口径；消除 Excel 中不可见空格、全角空格和换行差异。
    return re.sub(r"\s+", "", txt(v).replace("\u3000", ""))
def county(v):
    x=txt(v);return COUNTY_UPDATES.get(x,x)
def num(v):
    if v is None: return None
    m=re.search(r"[-+]?\d+(?:\.\d+)?",txt(v).replace(",",""))
    return float(m.group()) if m else None
def branch_amount(v):
    n=num(v)
    if n is None:return None
    s=txt(v); rate=1
    for k,c in [("美元","USD"),("港元","HKD"),("港币","HKD"),("欧元","EUR"),("德国马克","DEM")]:
        if k in s: rate=FX[c];break
    return n*(1 if "万" in s or isinstance(v,(int,float)) else .0001)*rate
def inv_amount(v,c):
    n=num(v); code=txt(c).upper() or "CNY"
    return None if n is None else n*FX.get(code,1)
def canonical_code(v,level):
    s=txt(v).upper();return s[:{1:1,2:3,3:4}[level]] if s else ""
def load_catalog():
    wb=load_workbook(INDUSTRY_CODE_FILE,read_only=True,data_only=True);ws=wb.active;rows=ws.iter_rows(values_only=True);h=next(rows);p={x:i for i,x in enumerate(h)};out={1:{},2:{},3:{}}
    for r in rows:
        try:level=int(r[p["series"]])
        except (TypeError,ValueError):continue
        if level in out:
            code=canonical_code(r[p["industry_code"]],level);name=txt(r[p["industry_name"]])
            if code:out[level][code]=name
    wb.close();return out
def industry(r,p,side,level):
    code=canonical_code(r[p.get(f"{side}_industry_code{level}")] if f"{side}_industry_code{level}" in p else "",level);source=txt(r[p.get(f"{side}_industry_name{level}")]) if f"{side}_industry_name{level}" in p else ""
    if not code:return "",source
    if code not in INDUSTRY_CATALOG[level]:return "","未分类（代码表无匹配）"
    return code,INDUSTRY_CATALOG[level][code]

def process(kind,path):
    wb=load_workbook(path,read_only=True,data_only=True); ws=wb.active
    rows=ws.iter_rows(values_only=True); h=next(rows); p={x:i for i,x in enumerate(h)}
    agg=defaultdict(lambda:[0,0.0,0]);town_agg=defaultdict(lambda:[0,0.0,0]);seen=set()
    for r in rows:
        oc,dc=txt(r[p["o_city"]]),txt(r[p["d_city"]])
        oo,dd=county(r[p["o_county"]]),county(r[p["d_county"]]);ot,dt=txt(r[p["o_village"]]),txt(r[p["d_village"]])
        if oc not in CITIES or dc not in CITIES or not oo or not dd: continue
        if kind=="投资": key=(txt(r[p["eid"]]) or txt(r[p["name"]]),txt(r[p["invest_eid"]]) or txt(r[p["invest_name"]]))
        else: key=(txt(r[p["eid"]]) or txt(r[p["o_name"]]),txt(r[p["sub_eid"]]) or txt(r[p["d_name"]]))
        if key in seen:continue
        seen.add(key)
        amount=inv_amount(r[p["should_capi_conv"]],r[p["currency_code"]]) if kind=="投资" else branch_amount(r[p["d_regist_capi"]])
        for level in (1,2,3):
            code,name=industry(r,p,"d",level)
            if (oc,oo)!=(dc,dd):
                k=(kind,level,code,name,oc,oo,dc,dd)
                agg[k][0]+=1
                if amount is not None: agg[k][1]+=amount;agg[k][2]+=1
            if ot and dt and (oc,oo,ot)!=(dc,dd,dt):
                tk=(kind,level,code,name,oc,oo,ot,dc,dd,dt)
                town_agg[tk][0]+=1
                if amount is not None:town_agg[tk][1]+=amount;town_agg[tk][2]+=1
    wb.close();return agg,town_agg

def patent_endpoint(r,p,side,entity):
    return {"entity":entity,"city":txt(r[p[f"{side}_city"]]),"county":county(r[p[f"{side}_county"]]),"town":txt(r[p[f"{side}_village"]]),"industries":{level:industry(r,p,side,level) for level in (1,2,3)}}

def process_patents(path):
    wb=load_workbook(path,read_only=True,data_only=True);ws=wb.active
    rows=ws.iter_rows(values_only=True);h=next(rows);p={x:i for i,x in enumerate(h)}
    grouped={};source_rows=0;eligible_rows=0;empty_name_rows=0;invalid_pair_rows=0
    for r in rows:
        source_rows+=1
        oc,dc=txt(r[p["o_city"]]),txt(r[p["d_city"]]);oo,dd=county(r[p["o_county"]]),county(r[p["d_county"]])
        if oc not in CITIES or dc not in CITIES or not oo or not dd:continue
        eligible_rows+=1
        patent=normalized_patent_name(r[p["patent_name"]])
        if not patent:
            empty_name_rows+=1
            continue
        oe=txt(r[p["o_eid"]]) or txt(r[p["o_name"]]);de=txt(r[p["d_eid"]]) or txt(r[p["d_name"]])
        if not oe or not de or oe==de:
            invalid_pair_rows+=1
            continue
        # role、记录方向、专利 ID 和申请号均不进入唯一键。
        pair=tuple(sorted((oe,de)))
        item=grouped.setdefault((patent,pair),{"directions":set(),"endpoints":{}});item["directions"].add((oe,de))
        for side,entity in (("o",oe),("d",de)):
            candidate=patent_endpoint(r,p,side,entity);previous=item["endpoints"].get(entity)
            score=sum(bool(c or n) for c,n in candidate["industries"].values())+4*bool(candidate["town"])
            old=-1 if previous is None else sum(bool(c or n) for c,n in previous["industries"].values())+4*bool(previous["town"])
            if score>old:item["endpoints"][entity]=candidate
    wb.close()
    county_meta=defaultdict(lambda:{"count":0,"patents":set(),"pairs":set(),"pair_patents":defaultdict(set),"unmatched":0})
    industry_meta={level:defaultdict(lambda:{"count":0,"patents":set(),"pairs":set(),"pair_patents":defaultdict(set),"unmatched":0}) for level in (1,2,3)}
    town_meta=defaultdict(lambda:{"count":0,"patents":set(),"pairs":set(),"pair_patents":defaultdict(set),"unmatched":0})
    town_industry_meta={level:defaultdict(lambda:{"count":0,"patents":set(),"pairs":set(),"pair_patents":defaultdict(set),"unmatched":0}) for level in (1,2,3)}
    for (patent,pair),item in grouped.items():
        if len(item["endpoints"])!=2:continue
        a,b=sorted(item["endpoints"].values(),key=lambda e:(e["city"],e["county"],e["entity"]));matched=(pair[0],pair[1]) in item["directions"] and (pair[1],pair[0]) in item["directions"]
        county_key=(a["city"],a["county"],b["city"],b["county"])
        county_distinct=county_key[:2]!=county_key[2:]
        if county_distinct:
            cm=county_meta[county_key];cm["count"]+=1;cm["patents"].add(patent);cm["pairs"].add(pair);cm["pair_patents"][pair].add(patent);cm["unmatched"]+=0 if matched else 1
        town_key=(a["city"],a["county"],a["town"],b["city"],b["county"],b["town"])
        town_distinct=town_key[:3]!=town_key[3:]
        if a["town"] and b["town"] and town_distinct:
            tm=town_meta[town_key];tm["count"]+=1;tm["patents"].add(patent);tm["pairs"].add(pair);tm["pair_patents"][pair].add(patent);tm["unmatched"]+=0 if matched else 1
        for level in (1,2,3):
            # 单行业筛选：任一企业端属于该行业即计入；双方同属一个行业时只计一次。
            for code,name in set((a["industries"][level],b["industries"][level])):
                if not code:continue
                if county_distinct:
                    key=(level,code,name,*county_key);im=industry_meta[level][key];im["count"]+=1;im["patents"].add(patent);im["pairs"].add(pair);im["pair_patents"][pair].add(patent);im["unmatched"]+=0 if matched else 1
                if a["town"] and b["town"] and town_distinct:
                    tkey=(level,code,name,*town_key);tim=town_industry_meta[level][tkey];tim["count"]+=1;tim["patents"].add(patent);tim["pairs"].add(pair);tim["pair_patents"][pair].add(patent);tim["unmatched"]+=0 if matched else 1
    records=[]
    for (oc,oo,dc,dd),m in county_meta.items():
        max_pair=max((len(v) for v in m["pair_patents"].values()),default=0);share=max_pair/len(m["patents"]) if m["patents"] else 0
        records.append(["专利",0,"ALL","全部行业对",oc,oo,dc,dd,m["count"],0,0,len(m["patents"]),len(m["pairs"]),max_pair,round(share,4),m["unmatched"]])
    for level,items in industry_meta.items():
        for (lv,code,name,oc,oo,dc,dd),m in items.items():
            max_pair=max((len(v) for v in m["pair_patents"].values()),default=0);share=max_pair/len(m["patents"]) if m["patents"] else 0
            records.append(["专利",lv,code,name,oc,oo,dc,dd,m["count"],0,0,len(m["patents"]),len(m["pairs"]),max_pair,round(share,4),m["unmatched"]])
    town_records=[]
    for (oc,oo,ot,dc,dd,dt),m in town_meta.items():
        max_pair=max((len(v) for v in m["pair_patents"].values()),default=0);share=max_pair/len(m["patents"]) if m["patents"] else 0
        town_records.append(["专利",0,"ALL","全部行业",oc,oo,ot,dc,dd,dt,m["count"],0,0,len(m["patents"]),len(m["pairs"]),max_pair,round(share,4),m["unmatched"]])
    for level,items in town_industry_meta.items():
        for (lv,code,name,oc,oo,ot,dc,dd,dt),m in items.items():
            max_pair=max((len(v) for v in m["pair_patents"].values()),default=0);share=max_pair/len(m["patents"]) if m["patents"] else 0
            town_records.append(["专利",lv,code,name,oc,oo,ot,dc,dd,dt,m["count"],0,0,len(m["patents"]),len(m["pairs"]),max_pair,round(share,4),m["unmatched"]])
    stats={"sourceRows":source_rows,"eligibleRows":eligible_rows,"dedupedRelations":len(grouped),"emptyPatentNameRows":empty_name_rows,"invalidEnterprisePairRows":invalid_pair_rows}
    return records,town_records,stats

def gpkg_wkb(blob):
    flags=blob[3]; env=(flags>>1)&7; sizes={0:0,1:32,2:48,3:48,4:64}
    return memoryview(blob)[8+sizes[env]:]
def parse_geom(data):
    pos=0
    def read():
        nonlocal pos
        endian="<" if data[pos]==1 else ">";pos+=1
        typ=struct.unpack_from(endian+"I",data,pos)[0];pos+=4; typ%=1000
        if typ==3:
            n=struct.unpack_from(endian+"I",data,pos)[0];pos+=4;r=[]
            for _ in range(n):
                m=struct.unpack_from(endian+"I",data,pos)[0];pos+=4
                ring=[]
                for _ in range(m):
                    x,y=struct.unpack_from(endian+"dd",data,pos);pos+=16;ring.append([x,y])
                r.append(ring)
            return [r]
        if typ==6:
            n=struct.unpack_from(endian+"I",data,pos)[0];pos+=4;out=[]
            for _ in range(n):out.extend(read())
            return out
        raise ValueError(typ)
    return read()
def simplify(ring,tolerance=0.00008):
    if len(ring)<4:return ring
    closed=ring[0]==ring[-1];points=ring[:-1] if closed else ring
    if len(points)<3:return ring
    keep={0,len(points)-1};stack=[(0,len(points)-1)];tolerance_sq=tolerance*tolerance
    while stack:
        start,end=stack.pop();ax,ay=points[start];bx,by=points[end];dx,dy=bx-ax,by-ay;length_sq=dx*dx+dy*dy
        furthest=-1;max_distance_sq=0.0
        for index in range(start+1,end):
            px,py=points[index]
            if length_sq:
                cross=(px-ax)*dy-(py-ay)*dx;distance_sq=cross*cross/length_sq
            else:distance_sq=(px-ax)**2+(py-ay)**2
            if distance_sq>max_distance_sq:furthest,max_distance_sq=index,distance_sq
        if furthest>=0 and max_distance_sq>tolerance_sq:
            keep.add(furthest);stack.extend(((start,furthest),(furthest,end)))
    out=[points[index] for index in sorted(keep)]
    if closed:out.append(out[0])
    return out

def polygon_centroid(polygons):
    total_x = total_y = total_area = 0.0
    for polygon in polygons:
        for ring in polygon:
            ring_area = 0.0
            ring_x = ring_y = 0.0
            for i in range(len(ring) - 1):
                x1, y1 = ring[i]
                x2, y2 = ring[i + 1]
                cross = x1 * y2 - x2 * y1
                ring_area += cross
                ring_x += (x1 + x2) * cross
                ring_y += (y1 + y2) * cross
            ring_area /= 2.0
            if abs(ring_area) < 1e-12:
                continue
            abs_area = abs(ring_area)
            total_x += ring_x / (6.0 * ring_area) * abs_area
            total_y += ring_y / (6.0 * ring_area) * abs_area
            total_area += abs_area
    if total_area <= 0:
        return None
    return [total_x / total_area, total_y / total_area]


def load_town_boundaries(conn):
    tables={row[0] for row in conn.execute("select table_name from gpkg_contents")}
    if "厦漳泉乡镇街边界" not in tables:return []
    features=[]
    for city,county_name,town,blob in conn.execute('select [城市名称],[区县名称],[乡镇街名称],geom from [厦漳泉乡镇街边界]'):
        raw=parse_geom(gpkg_wkb(blob));points=[point for poly in raw for ring in poly for point in ring]
        if not points:continue
        xs=[point[0] for point in points];ys=[point[1] for point in points]
        center=polygon_centroid(raw) or [(min(xs)+max(xs))/2,(min(ys)+max(ys))/2]
        polys=[[simplify(r) for r in poly] for poly in raw]
        features.append({"type":"Feature","properties":{"city":txt(city),"county":county(county_name),"town":txt(town),"center":center},"geometry":{"type":"MultiPolygon","coordinates":polys}})
    return features

def main():
    global INDUSTRY_CATALOG
    INDUSTRY_CATALOG=load_catalog()
    OUT.mkdir(parents=True,exist_ok=True)
    specs=[("投资",SOURCE/"t_hu_investments_it_xzq.xlsx"),("分支",SOURCE/"t_hu_branches_it_xzq.xlsx"),("专利",SOURCE/"t_hu_patents_relations_sq_xzq.xlsx")]
    allagg=defaultdict(lambda:[0,0.0,0]);all_townagg=defaultdict(lambda:[0,0.0,0])
    for kind,path in specs[:2]:
        print("处理",kind,flush=True)
        county_result,town_result=process(kind,path)
        for k,v in county_result.items():
            allagg[k][0]+=v[0];allagg[k][1]+=v[1];allagg[k][2]+=v[2]
        for k,v in town_result.items():
            all_townagg[k][0]+=v[0];all_townagg[k][1]+=v[1];all_townagg[k][2]+=v[2]
    records=[]; industries={level:dict(INDUSTRY_CATALOG[level]) for level in (1,2,3)}
    for k,v in allagg.items():
        kind,level,code,name,oc,oo,dc,dd=k
        records.append([kind,level,code,name,oc,oo,dc,dd,v[0],round(v[1],4),v[2],0,0,0,0,0])
    print("处理 专利（无向去重）",flush=True)
    patent_records,patent_town_records,patent_stats=process_patents(specs[2][1]);records.extend(patent_records)
    town_records=[]
    for k,v in all_townagg.items():
        kind,level,code,name,oc,oo,ot,dc,dd,dt=k
        town_records.append([kind,level,code,name,oc,oo,ot,dc,dd,dt,v[0],round(v[1],4),v[2],0,0,0,0,0])
    town_records.extend(patent_town_records)
    if GPKG.exists():
        conn=sqlite3.connect(GPKG);features=[];centers={}
        for code,name,blob in conn.execute('select XZQDM,XZQMC,geom from [县级调查界限]'):
            code=txt(code)
            if code[:4] not in CITY_CODES:continue
            raw=parse_geom(gpkg_wkb(blob));points=[point for poly in raw for ring in poly for point in ring]
            if not points:continue
            county_name=county(name);xs=[point[0] for point in points];ys=[point[1] for point in points]
            centers[county_name]=[(min(xs)+max(xs))/2,(min(ys)+max(ys))/2]
            polys=[[simplify(r) for r in poly] for poly in raw]
            features.append({"type":"Feature","properties":{"name":county_name,"city":CITY_CODES[code[:4]],"code":code},"geometry":{"type":"MultiPolygon","coordinates":polys}})
        town_boundaries=load_town_boundaries(conn);conn.close()
    else:
        previous=json.loads((OUT/"enterprise-relations.json").read_text(encoding="utf-8"));features=previous["boundaries"]["features"];centers=previous["centers"]
        town_path=OUT/"township-boundaries.json";town_boundaries=json.loads(town_path.read_text(encoding="utf-8")).get("features",[]) if town_path.exists() else []
    payload={"meta":{"baseline":"2026-07-21","amountUnit":"万元人民币","sameOdExcluded":True,"patentUndirected":True,"patentDedupKey":"normalized_patent_name+unordered_enterprise_pair","patentIndustryFilter":"either_endpoint","patentDiagnostics":patent_stats,"sources":[x[1].name for x in specs]},"industries":{str(k):[[c,n] for c,n in sorted(v.items())] for k,v in industries.items()},"centers":centers,"boundaries":{"type":"FeatureCollection","features":features},"records":records}
    (OUT/"enterprise-relations.json").write_text(json.dumps(payload,ensure_ascii=False,separators=(",",":")),encoding="utf-8")
    (OUT/"township-boundaries.json").write_text(json.dumps({"type":"FeatureCollection","features":town_boundaries},ensure_ascii=False,separators=(",",":")),encoding="utf-8")
    town_pair_count=write_town_endpoints(town_records)
    patent_count=sum(r[8] for r in patent_records if r[1]==0)
    append_dev_log("专利去重改为专利名称口径",f"使用“标准化专利名＋无序企业对”合并正反向及 role 重复；不再使用专利 ID 或申请号。符合区县范围的 {patent_stats['eligibleRows']} 条源记录合并为 {patent_stats['dedupedRelations']} 条专利—企业对关系，跳过 {patent_stats['emptyPatentNameRows']} 条专利名为空记录。")
    append_dev_log("镇街关系接口已更新",f"生成 {len(town_records)} 个镇街汇总组合、{town_pair_count} 个区县对分片和 {len(town_boundaries)} 个乡镇街边界；仅统计 O/D 两端镇街字段均有效的记录。")
    print("输出",OUT/"enterprise-relations.json",(OUT/"enterprise-relations.json").stat().st_size)
    print("输出",OUT/"township-index.json",town_pair_count,"个区县对分片")
    print("输出",OUT/"township-relations.json",len(town_records),"条全量镇街汇总")
if __name__=="__main__":main()
