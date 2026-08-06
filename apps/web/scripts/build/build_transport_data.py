from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PLATFORM_ROOT = Path(__file__).resolve().parents[4]
SOURCE = PLATFORM_ROOT / "data" / "raw" / "transport" / "driving" / "county-government" / "厦漳泉区县政府所在地_OD.xlsx"
TOWNSHIP_SOURCE = PLATFORM_ROOT / "data" / "raw" / "transport" / "driving" / "township-government" / "厦漳泉乡镇街政府经纬度_WGS84_OD_驾车结果.xlsx"
POPULATION_INTERFACE = Path(__file__).resolve().parents[2] / "public" / "data" / "population-flow.json"
TOWNSHIP_BOUNDARIES = Path(__file__).resolve().parents[2] / "public" / "data" / "township-boundaries.json"
OUT = Path(__file__).resolve().parents[2] / "public" / "data" / "transport-accessibility.json"
TOWNSHIP_OUT = Path(__file__).resolve().parents[2] / "public" / "data" / "transport-accessibility-township.json"
CENTER_OUT = Path(__file__).resolve().parents[2] / "public" / "data" / "government-centers.json"


def avg(values):
    values = [value for value in values if value is not None]
    return sum(values) / len(values) if values else 0.0


def build_township(population):
    workbook = load_workbook(TOWNSHIP_SOURCE, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = {name: position for position, name in enumerate(headers)}
    required = [
        "O_城市名称", "O_区县名称", "O_乡镇街名称", "O_wgs84_经", "O_wgs84_纬",
        "D_城市名称", "D_区县名称", "D_乡镇街名称", "D_wgs84_经", "D_wgs84_纬",
        "驾车时间_秒", "驾车距离_米", "过路费_元", "高德查询状态",
    ]
    missing = [name for name in required if name not in index]
    if missing:
        raise ValueError(f"乡镇街数据缺少字段: {missing}")

    records, centers, by_direction = [], {}, {}
    node_outbound = defaultdict(list)
    failed_rows = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        origin = tuple(str(values[index[f"O_{field}"]] or "").strip() for field in ("城市名称", "区县名称", "乡镇街名称"))
        destination = tuple(str(values[index[f"D_{field}"]] or "").strip() for field in ("城市名称", "区县名称", "乡镇街名称"))
        status = str(values[index["高德查询状态"]] or "").strip()
        if status != "成功":
            failed_rows.append(row_number)
            continue
        if not all(origin) or not all(destination) or origin == destination:
            continue
        duration_seconds = float(values[index["驾车时间_秒"]])
        distance_meters = float(values[index["驾车距离_米"]])
        toll = float(values[index["过路费_元"]])
        item = [*origin, *destination, duration_seconds, duration_seconds / 60, distance_meters, distance_meters / 1000, toll]
        direction = (*origin, *destination)
        if direction in by_direction:
            raise ValueError(f"乡镇街数据存在重复有向OD: {direction}")
        records.append(item)
        by_direction[direction] = item
        node_outbound[origin].append(item)
        centers["|".join(origin)] = [float(values[index["O_wgs84_经"]]), float(values[index["O_wgs84_纬"]])]
        centers["|".join(destination)] = [float(values[index["D_wgs84_经"]]), float(values[index["D_wgs84_纬"]])]
    workbook.close()

    nodes = sorted(centers)
    node_tuples = [tuple(node.split("|")) for node in nodes]
    incomplete_pairs, pair_records = [], []
    for left_pos, left in enumerate(node_tuples):
        for right in node_tuples[left_pos + 1:]:
            forward = by_direction.get((*left, *right))
            backward = by_direction.get((*right, *left))
            if not forward or not backward:
                incomplete_pairs.append([*left, *right])
                continue
            pair_records.append([*left, *right, avg([forward[7], backward[7]]), avg([forward[9], backward[9]]), avg([forward[10], backward[10]]), forward[7], backward[7], forward[9], backward[9], forward[10], backward[10]])

    node_stats = []
    for node in node_tuples:
        items = node_outbound[node]
        node_stats.append({
            "city": node[0], "county": node[1], "town": node[2],
            "avgTime": avg([row[7] for row in items]),
            "avgDistance": avg([row[9] for row in items]),
            "avgToll": avg([row[10] for row in items]),
            "coverage60": sum(row[7] <= 60 for row in items),
            "coverage90": sum(row[7] <= 90 for row in items),
        })
    node_stats.sort(key=lambda item: item["avgTime"])
    for rank, item in enumerate(node_stats, start=1):
        item["rank"] = rank

    boundary_payload = json.loads(TOWNSHIP_BOUNDARIES.read_text(encoding="utf-8"))
    township_features = boundary_payload["features"]
    payload = {
        "meta": {
            "source": TOWNSHIP_SOURCE.name,
            "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
            "level": "镇街", "nodeCount": len(nodes), "directedOdCount": len(records),
            "expectedDirectedOdCount": len(nodes) * (len(nodes) - 1),
            "pairCount": len(pair_records), "incompletePairCount": len(incomplete_pairs),
            "boundaryCount": len(township_features), "odNodeBoundaryCount": len(centers),
            "failedRowCount": len(failed_rows), "durationUnit": "分钟", "distanceUnit": "公里", "tollUnit": "元",
            "distanceDefinition": "无向镇街对距离取A→B与B→A政府驻地驾车里程的算术平均值" + ("；缺少任一方向时不生成无向均值" if incomplete_pairs else "；全部镇街对均具有双向记录"),
        },
        "governmentCenters": centers,
        "countyBoundaries": population["countyBoundaries"],
        "townshipBoundaries": {"type": "FeatureCollection", "features": township_features},
        "records": records, "pairRecords": pair_records, "nodeStats": node_stats,
        "incompletePairs": incomplete_pairs,
    }
    TOWNSHIP_OUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    return payload


def main():
    population = json.loads(POPULATION_INTERFACE.read_text(encoding="utf-8"))
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = {name: position for position, name in enumerate(headers)}
    required = [
        "O_cityname", "O_adname", "O_wgs84_经", "O_wgs84_纬",
        "D_cityname", "D_adname", "D_wgs84_经", "D_wgs84_纬",
        "耗时", "里程（米）", "过路费（元）",
    ]
    missing = [name for name in required if name not in index]
    if missing:
        raise ValueError(f"缺少字段: {missing}")

    records = []
    centers = {}
    by_direction = {}
    node_outbound = defaultdict(list)
    for values in sheet.iter_rows(min_row=2, values_only=True):
        oc, county_o = str(values[index["O_cityname"]] or "").strip(), str(values[index["O_adname"]] or "").strip()
        dc, county_d = str(values[index["D_cityname"]] or "").strip(), str(values[index["D_adname"]] or "").strip()
        if not oc or not county_o or not dc or not county_d or (oc == dc and county_o == county_d):
            continue
        duration_seconds = float(values[index["耗时"]] or 0)
        distance_meters = float(values[index["里程（米）"]] or 0)
        toll = float(values[index["过路费（元）"]] or 0)
        ox, oy = float(values[index["O_wgs84_经"]]), float(values[index["O_wgs84_纬"]])
        dx, dy = float(values[index["D_wgs84_经"]]), float(values[index["D_wgs84_纬"]])
        item = [oc, county_o, dc, county_d, duration_seconds, duration_seconds / 60, distance_meters, distance_meters / 1000, toll]
        records.append(item)
        by_direction[(oc, county_o, dc, county_d)] = item
        node_outbound[(oc, county_o)].append(item)
        centers[f"{oc}|{county_o}"] = [ox, oy]
        centers[f"{dc}|{county_d}"] = [dx, dy]
    workbook.close()

    nodes = sorted(node_outbound)
    pair_records = []
    incomplete_pairs = []
    for left_pos, left in enumerate(nodes):
        for right in nodes[left_pos + 1:]:
            forward = by_direction.get((*left, *right))
            backward = by_direction.get((*right, *left))
            if not forward or not backward:
                incomplete_pairs.append([*left, *right])
                continue
            pair_records.append([
                *left, *right,
                avg([forward[5], backward[5]]),
                avg([forward[7], backward[7]]),
                avg([forward[8], backward[8]]),
                forward[5], backward[5], forward[7], backward[7], forward[8], backward[8],
            ])

    node_stats = []
    for city, county in nodes:
        items = node_outbound[(city, county)]
        node_stats.append({
            "city": city,
            "county": county,
            "avgTime": avg([row[5] for row in items]),
            "avgDistance": avg([row[7] for row in items]),
            "avgToll": avg([row[8] for row in items]),
            "coverage90": sum(row[5] <= 90 for row in items),
            "coverage120": sum(row[5] <= 120 for row in items),
        })
    node_stats.sort(key=lambda item: item["avgTime"])
    for rank, item in enumerate(node_stats, start=1):
        item["rank"] = rank

    city_groups = defaultdict(list)
    for row in records:
        city_groups[(row[0], row[2])].append(row)
    city_pair_stats = [{
        "originCity": key[0],
        "destinationCity": key[1],
        "avgTime": avg([row[5] for row in items]),
        "avgDistance": avg([row[7] for row in items]),
        "avgToll": avg([row[8] for row in items]),
        "odCount": len(items),
    } for key, items in sorted(city_groups.items())]

    payload = {
        "meta": {
            "source": SOURCE.name,
            "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
            "nodeCount": len(nodes),
            "directedOdCount": len(records),
            "pairCount": len(pair_records),
            "incompletePairCount": len(incomplete_pairs),
            "durationUnit": "分钟",
            "distanceUnit": "公里",
            "tollUnit": "元",
            "distanceDefinition": "无向区县对距离取A→B与B→A区县政府驻地驾车里程的算术平均值",
        },
        "governmentCenters": centers,
        "countyBoundaries": population["countyBoundaries"],
        "records": records,
        "pairRecords": pair_records,
        "nodeStats": node_stats,
        "cityPairStats": city_pair_stats,
        "incompletePairs": incomplete_pairs,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    township_payload = build_township(population)
    CENTER_OUT.write_text(json.dumps({
        "meta": {"generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"), "crs": "WGS84", "countyCount": len(centers), "townshipCount": len(township_payload["governmentCenters"])},
        "county": centers,
        "township": township_payload["governmentCenters"],
    }, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(json.dumps({"county": {"output": str(OUT), **payload["meta"]}, "township": {"output": str(TOWNSHIP_OUT), **township_payload["meta"]}}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
