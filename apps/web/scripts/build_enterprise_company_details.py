from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from build_data import CITIES, SOURCE, OUT, branch_amount, county, industry, inv_amount, load_catalog, txt
import build_data


def pair_key(kind: str, oc: str, oo: str, dc: str, dd: str) -> str:
    return "|".join((kind, oc, oo, dc, dd))


def collect(kind: str, path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    positions = {name: index for index, name in enumerate(header)}
    grouped = defaultdict(list)
    seen = set()
    for row in rows:
        oc, dc = txt(row[positions["o_city"]]), txt(row[positions["d_city"]])
        oo, dd = county(row[positions["o_county"]]), county(row[positions["d_county"]])
        if oc not in CITIES or dc not in CITIES or not oo or not dd or (oc, oo) == (dc, dd):
            continue
        if kind == "投资":
            o_name, d_name = txt(row[positions["name"]]), txt(row[positions["invest_name"]])
            unique = (txt(row[positions["eid"]]) or o_name, txt(row[positions["invest_eid"]]) or d_name)
            amount = inv_amount(row[positions["should_capi_conv"]], row[positions["currency_code"]])
        else:
            o_name, d_name = txt(row[positions["o_name"]]), txt(row[positions["d_name"]])
            unique = (txt(row[positions["eid"]]) or o_name, txt(row[positions["sub_eid"]]) or d_name)
            amount = branch_amount(row[positions["d_regist_capi"]])
        if unique in seen:
            continue
        seen.add(unique)
        industries = [industry(row, positions, "d", level) for level in (1, 2, 3)]
        grouped[pair_key(kind, oc, oo, dc, dd)].append([
            o_name or unique[0], d_name or unique[1], amount,
            industries[0][0], industries[0][1], industries[1][0], industries[1][1], industries[2][0], industries[2][1]
        ])
    workbook.close()
    return grouped


def main():
    build_data.INDUSTRY_CATALOG = load_catalog()
    target = OUT / "enterprise-companies"
    target.mkdir(parents=True, exist_ok=True)
    for old in target.glob("pair-*.json"):
        old.unlink()
    grouped = defaultdict(list)
    for kind, filename in (("投资", "t_hu_investments_it_xzq.xlsx"), ("分支", "t_hu_branches_it_xzq.xlsx")):
        print("整理企业明细", kind, flush=True)
        for key, records in collect(kind, SOURCE / filename).items():
            grouped[key].extend(records)
    index = {}
    for number, key in enumerate(sorted(grouped), 1):
        filename = f"pair-{number:04d}.json"
        relation = key.split("|", 1)[0]
        payload = {
            "meta": {
                "relation": relation,
                "amountLabel": "投资额" if relation == "投资" else "分支注册资本",
                "amountUnit": "万元人民币",
                "fields": ["oCompany", "dCompany", "amount", "industry1Code", "industry1Name", "industry2Code", "industry2Name", "industry3Code", "industry3Name"],
            },
            "records": grouped[key],
        }
        (target / filename).write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        index[key] = filename
    (target / "index.json").write_text(json.dumps({"pairs": index, "pairCount": len(index)}, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print("输出企业明细分片", len(index), "个区县 OD", flush=True)


if __name__ == "__main__":
    main()
