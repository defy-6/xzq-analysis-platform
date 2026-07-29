from __future__ import annotations

import json
import sqlite3
import struct
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PLATFORM_ROOT = Path(__file__).resolve().parents[3]
POPULATION_DIR = PLATFORM_ROOT / "data" / "raw" / "population"
SOURCE_FILE = POPULATION_DIR / "厦漳泉乡镇级人口流动_汇总版.xlsx"
GPKG = PLATFORM_ROOT / "data" / "raw" / "spatial" / "fujian.gpkg"
OUT = Path(__file__).resolve().parents[1] / "public" / "data" / "population-flow.json"
CITY_CODES = {"3502": "厦门市", "3505": "泉州市", "3506": "漳州市"}
CITIES = set(CITY_CODES.values())


def text(value):
    return "" if value is None else str(value).strip()


def normalize_town(value):
    name = text(value)
    for suffix in ("街道办事处", "街道", "民族乡", "镇", "乡"):
        if name.endswith(suffix):
            return name[: -len(suffix)]
    return name


def gpkg_wkb(blob):
    flags = blob[3]
    envelope = (flags >> 1) & 7
    sizes = {0: 0, 1: 32, 2: 48, 3: 48, 4: 64}
    return memoryview(blob)[8 + sizes[envelope] :]


def parse_geometry(data):
    position = 0

    def read():
        nonlocal position
        endian = "<" if data[position] == 1 else ">"
        position += 1
        geometry_type = struct.unpack_from(endian + "I", data, position)[0] % 1000
        position += 4
        if geometry_type == 3:
            ring_count = struct.unpack_from(endian + "I", data, position)[0]
            position += 4
            rings = []
            for _ in range(ring_count):
                point_count = struct.unpack_from(endian + "I", data, position)[0]
                position += 4
                ring = []
                for _ in range(point_count):
                    x, y = struct.unpack_from(endian + "dd", data, position)
                    position += 16
                    ring.append([x, y])
                rings.append(ring)
            return [rings]
        if geometry_type == 6:
            polygon_count = struct.unpack_from(endian + "I", data, position)[0]
            position += 4
            polygons = []
            for _ in range(polygon_count):
                polygons.extend(read())
            return polygons
        raise ValueError(f"Unsupported geometry type: {geometry_type}")

    return read()


def simplify(ring, target_points=300):
    if len(ring) <= target_points:
        return ring
    step = max(2, len(ring) // target_points)
    result = ring[::step]
    if result[-1] != ring[-1]:
        result.append(ring[-1])
    return result


def bounds(polygons):
    points = [point for polygon in polygons for ring in polygon for point in ring]
    xs = [point[0] for point in points]
    ys = [point[1] for point in points]
    return min(xs), min(ys), max(xs), max(ys)


def load_boundaries(connection):
    towns = []
    query = """
        select [城市名称], [区县名称], [乡镇街名称], Shape_Area, geom
        from [厦漳泉乡镇街边界]
    """
    for city, county, town, area, blob in connection.execute(query):
        polygons = parse_geometry(gpkg_wkb(blob))
        bbox = bounds(polygons)
        towns.append(
            {
                "city": text(city),
                "county": text(county),
                "town": text(town),
                "center": [(bbox[0] + bbox[2]) / 2, (bbox[1] + bbox[3]) / 2],
            }
        )

    county_features = []
    county_centers = {}
    county_query = "select XZQDM, XZQMC, geom from [县级调查界限]"
    for code, name, blob in connection.execute(county_query):
        code = text(code)
        if code[:4] not in CITY_CODES:
            continue
        polygons = parse_geometry(gpkg_wkb(blob))
        bbox = bounds(polygons)
        county_name = text(name)
        city = CITY_CODES[code[:4]]
        county_centers[f"{city}|{county_name}"] = [(bbox[0] + bbox[2]) / 2, (bbox[1] + bbox[3]) / 2]
        county_features.append(
            {
                "type": "Feature",
                "properties": {"city": city, "name": county_name, "code": code},
                "geometry": {
                    "type": "MultiPolygon",
                    "coordinates": [[simplify(ring) for ring in polygon] for polygon in polygons],
                },
            }
        )
    return towns, county_features, county_centers


def main():
    with sqlite3.connect(GPKG) as connection:
        towns, county_features, county_centers = load_boundaries(connection)

    workbook = load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    column = {name: index for index, name in enumerate(headers)}

    county_aggregate = defaultdict(lambda: [0, 0])
    town_aggregate = defaultdict(lambda: [0, 0])
    town_centers = {}
    boundary_town_centers = {}
    normalized_boundary_town_centers = {}
    for town in towns:
        exact_key = f'{town["city"]}|{town["county"]}|{town["town"]}'
        normalized_key = f'{town["city"]}|{town["county"]}|~{normalize_town(town["town"])}'
        boundary_town_centers[exact_key] = town["center"]
        normalized_boundary_town_centers.setdefault(normalized_key, town["center"])
    raw_rows = positive_rows = valid_rows = 0
    excluded_nonpositive = excluded_incomplete = 0
    excluded_same_town = excluded_same_town_population = 0
    fallback_town_names = set()
    within_county_population = 0

    endpoints = (
        ("起点地市名称", "起点区县名称", "起点乡镇名称"),
        ("终到地市名称", "终到区县名称", "终到乡镇名称"),
    )

    for row in rows:
        raw_rows += 1
        population = int(row[column["人口数量"]] or 0)
        if population <= 0:
            excluded_nonpositive += 1
            continue
        positive_rows += 1
        source_endpoints = [
            (text(row[column[city_column]]), text(row[column[county_column]]), text(row[column[town_column]]))
            for city_column, county_column, town_column in endpoints
        ]
        if any(city not in CITIES or not county or not town for city, county, town in source_endpoints):
            excluded_incomplete += 1
            continue
        origin_tuple, destination_tuple = source_endpoints
        if origin_tuple == destination_tuple:
            excluded_same_town += 1
            excluded_same_town_population += population
            continue
        valid_rows += 1
        for city, county, town in source_endpoints:
            key = f"{city}|{county}|{town}"
            if key in town_centers:
                continue
            normalized_key = f"{city}|{county}|~{normalize_town(town)}"
            center = boundary_town_centers.get(key) or normalized_boundary_town_centers.get(normalized_key)
            if center is None:
                center = county_centers.get(f"{city}|{county}")
                fallback_town_names.add(key)
            if center is not None:
                town_centers[key] = center

        town_key = (*origin_tuple, *destination_tuple)
        town_aggregate[town_key][0] += population
        town_aggregate[town_key][1] += 1

        if origin_tuple[:2] == destination_tuple[:2]:
            within_county_population += population
            continue
        county_key = (*origin_tuple[:2], *destination_tuple[:2])
        county_aggregate[county_key][0] += population
        county_aggregate[county_key][1] += 1

    workbook.close()

    county_records = [
        [*key, values[0], values[1]] for key, values in county_aggregate.items() if values[0] > 0
    ]
    town_records = [
        [*key, values[0], values[1]] for key, values in town_aggregate.items() if values[0] > 0
    ]
    county_records.sort(key=lambda item: item[4], reverse=True)
    town_records.sort(key=lambda item: item[6], reverse=True)

    payload = {
        "meta": {
            "source": SOURCE_FILE.name,
            "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
            "unit": "人",
            "rawRows": raw_rows,
            "positiveRows": positive_rows,
            "validRows": valid_rows,
            "excludedNonPositive": excluded_nonpositive,
            "excludedIncomplete": excluded_incomplete,
            "excludedSameTown": excluded_same_town,
            "excludedSameTownPopulation": excluded_same_town_population,
            "fallbackTownNames": len(fallback_town_names),
            "withinCountyPopulation": within_county_population,
            "rule": "行政区代码不参与判断，地市、区县和镇街均直接使用源文件名称；排除人口数量不大于0、两端名称不完整或O/D为同一镇街的记录。",
        },
        "countyCenters": county_centers,
        "townCenters": town_centers,
        "countyBoundaries": {"type": "FeatureCollection", "features": county_features},
        "countyRecords": county_records,
        "townRecords": town_records,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(
        json.dumps(
            {
                "output": str(OUT),
                "size": OUT.stat().st_size,
                "countyRecords": len(county_records),
                "townRecords": len(town_records),
                **payload["meta"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
